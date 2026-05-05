"""
cargar_datos.py  — Convierte plantillas Excel → js/data.js
============================================================
Uso:  python cargar_datos.py
Req:  pip install openpyxl
"""
import os, re
from datetime import date, datetime
from openpyxl import load_workbook

BASE   = os.path.dirname(os.path.abspath(__file__))
PLANT  = os.path.join(BASE, "plantillas")
OUT_JS = os.path.join(BASE, "js", "data.js")

# ── helpers ───────────────────────────────────────────────────────────────────

def read_sheet(path, sheet_name):
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        d = {}
        for h, v in zip(headers, row):
            if isinstance(v, (date, datetime)):
                v = v.strftime("%Y-%m-%d") if isinstance(v, datetime) else v.isoformat()
            d[h] = v
        result.append(d)
    return result

def col(row, *keys, default=None):
    for k in keys:
        v = row.get(k)
        if v is not None and str(v).strip() != "" and str(v).strip() != "None":
            return v
    return default

def to_int(v, default=0):
    try:
        return int(float(str(v).replace(",", "").replace("$", "").strip()))
    except:
        return default

def to_float(v, default=0.0):
    try:
        return float(str(v).replace(",", "").replace("$", "").strip())
    except:
        return default

def s(v):
    """JS string literal."""
    if v is None:
        return '""'
    v = str(v).replace("\\", "\\\\").replace('"', '\\"').replace("\n", " ")
    return f'"{v}"'

def n(v):
    """JS number literal."""
    if v is None:
        return "0"
    try:
        f = float(str(v).replace(",", "").replace("$", ""))
        return str(int(f)) if f == int(f) else str(round(f, 4))
    except:
        return "0"

# ── read static header from current data.js ──────────────────────────────────

STATIC_END_MARKERS = [
    r'\n// ─── Datos generados',
    r'\nconst PDM\s*=',
    r'\nlet PROYECTOS\s*=',
]

def read_static_header():
    try:
        with open(OUT_JS, "r", encoding="utf-8") as f:
            content = f.read()
        # Find start of generated section
        for marker in STATIC_END_MARKERS:
            m = re.search(marker, content)
            if m:
                return content[:m.start()].rstrip()
        # Fallback: keep everything up to PROYECTOS comment
        m = re.search(r'\n// -+\n// PROYECTOS', content)
        if m:
            return content[:m.start()].rstrip()
        return content
    except:
        return "// data.js generado por cargar_datos.py"

# ── PDM ───────────────────────────────────────────────────────────────────────

def build_pdm():
    path = os.path.join(PLANT, "01_PDM.xlsx")
    if not os.path.exists(path):
        print("  [AVISO] 01_PDM.xlsx no encontrado")
        return 'const PDM = {municipio:"",periodo:"2024-2027",nombre:"",alcalde:"",acuerdo:"",inversionTotal:0,ejes:[]};'

    wb = load_workbook(path, data_only=True)
    ws = wb["InfoGeneral"]
    info = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1] is not None:
            info[str(row[0]).strip()] = row[1]

    ejes_raw = read_sheet(path, "EjesProgramas")
    ejes_dict = {}
    for r in ejes_raw:
        eid  = str(col(r, "ID Eje") or "EJE1")
        # Ignorar filas de notas/advertencias del template
        if eid.startswith("⚠") or eid.startswith("*"):
            continue
        enom = str(col(r, "Nombre Eje") or eid)
        pid  = str(col(r, "ID Programa") or "PROG")
        pnom = str(col(r, "Nombre Programa") or pid)
        lb   = to_float(col(r, "Línea Base", default=0))
        presup = to_int(col(r, "Presupuesto Programa (COP)", default=0))
        meta   = str(col(r, "Meta Cuatrienio") or "")
        indic  = str(col(r, "Indicador Resultado") or "")
        if eid not in ejes_dict:
            ejes_dict[eid] = {"id": eid, "nombre": enom, "programas": []}
        ejes_dict[eid]["programas"].append(
            f'{{id:{s(pid)},nombre:{s(pnom)},meta:{s(meta)},indicador:{s(indic)},lineaBase:{lb},presupuesto:{presup}}}'
        )

    ejes_js = []
    for eje in ejes_dict.values():
        progs = ",".join(eje["programas"])
        ejes_js.append(f'{{id:{s(eje["id"])},nombre:{s(eje["nombre"])},programas:[{progs}]}}')

    # Leer IndicadoresRadar desde 05_IndicadoresPDM.xlsx si existe
    radar_path = os.path.join(PLANT, "05_IndicadoresPDM.xlsx")
    radar_js = "[]"
    if os.path.exists(radar_path):
        radar_rows = read_sheet(radar_path, "IndicadoresRadar")
        radar_items = []
        for r in radar_rows:
            eid_r = str(col(r, "ID Eje") or "")
            if not eid_r or eid_r.startswith("⚠"):
                continue
            eje_label = str(col(r, "Nombre Eje (etiqueta corta)") or eid_r)
            logrado   = to_float(col(r, "Logrado (%)", default=0))
            meta_r    = to_float(col(r, "Meta (%)", default=100))
            radar_items.append(f'{{eje:{s(eje_label)},logrado:{logrado},meta:{meta_r}}}')
        if radar_items:
            radar_js = "[" + ",".join(radar_items) + "]"
            print(f"  IndicadoresRadar: {len(radar_items)} ejes cargados desde 05_IndicadoresPDM.xlsx")

    inversion = to_int(info.get("Inversión Total PDM (COP)", 85000000000))
    return (
        f'const PDM = {{\n'
        f'  municipio:{s(str(info.get("Municipio","Municipio")))},\n'
        f'  periodo:{s(str(info.get("Período","2024-2027")))},\n'
        f'  nombre:{s(str(info.get("Nombre del PDM","")))},\n'
        f'  alcalde:{s(str(info.get("Alcalde","")))} ,\n'
        f'  acuerdo:{s(str(info.get("Acuerdo Municipal","")))},\n'
        f'  inversionTotal:{inversion},\n'
        f'  indicadoresRadar:{radar_js},\n'
        f'  ejes:[{",".join(ejes_js)}]\n'
        f'}};'
    )

# ── PROYECTOS ─────────────────────────────────────────────────────────────────

def build_proyectos():
    path = os.path.join(PLANT, "02_Proyectos.xlsx")
    if not os.path.exists(path):
        print("  [AVISO] 02_Proyectos.xlsx no encontrado")
        return "let PROYECTOS = [];"

    proyectos  = read_sheet(path, "Proyectos")
    ejecucion  = read_sheet(path, "EjecucionPresupuestal")
    hitos_all  = read_sheet(path, "Hitos")
    contratos_all = read_sheet(path, "Contratos")
    print(f"  Proyectos: {len(proyectos)} | Ejecución: {len(ejecucion)} | Hitos: {len(hitos_all)} | Contratos: {len(contratos_all)}")

    items = []
    for p in proyectos:
        bpin = str(col(p, "BPIN") or "")
        nombre  = str(col(p, "Nombre Proyecto") or f"Proyecto {bpin}")
        sector  = str(col(p, "Sector") or "GOBIERNO").upper()
        estado  = str(col(p, "Estado") or "EN_EJECUCION").upper()
        prog_id = str(col(p, "ID Programa PDM") or "")
        desc    = str(col(p, "Descripcion") or "")
        obj_g   = str(col(p, "Objetivo General") or "")
        fecha_i = str(col(p, "Fecha Inicio") or "")
        fecha_f = str(col(p, "Fecha Fin") or "")
        pob_ben = to_int(col(p, "Población Beneficiada", default=0))
        pob_tip = str(col(p, "Tipo Población") or "")
        resp    = str(col(p, "Responsable") or "")
        obs     = str(col(p, "Observaciones") or "")
        cod_dnp = str(col(p, "Código Producto DNP", "Codigo Producto DNP") or "")
        ind_dnp = str(col(p, "Indicador DNP") or "")
        uni_dnp = str(col(p, "Unidad DNP") or "")

        # Ejecucion por vigencia
        ejec_p = [e for e in ejecucion if str(col(e, "BPIN") or "") == bpin]
        ejec_js_list = []
        total_aprop = 0
        total_pagos = 0
        for e in ejec_p:
            vig  = to_int(col(e, "Vigencia", default=0))
            apro = to_int(col(e, "Apropiacion Final (COP)", "Apropiacion Inicial (COP)", default=0))
            cdp  = to_int(col(e, "CDP (COP)", default=0))
            rp   = to_int(col(e, "RP (COP)", default=0))
            obl  = to_int(col(e, "Obligaciones (COP)", default=0))
            pag  = to_int(col(e, "Pagos (COP)", default=0))
            fuente_e = str(col(e, "Fuente", default=""))
            ejec_js_list.append(
                f'{{vigencia:{vig},apropiacion:{apro},cdp:{cdp},rp:{rp},obligaciones:{obl},pagos:{pag},fuente:{s(fuente_e)}}}'
            )
            total_aprop += apro
            total_pagos += pag

        ejec_js = "[" + ",".join(ejec_js_list) + "]"

        # valorTotal: suma de apropiacion de todas las vigencias, o fuentes
        valor_total = total_aprop if total_aprop > 0 else to_int(col(p, "valorTotal", default=0))

        # Avances
        hitos_p = [h for h in hitos_all if str(col(h, "BPIN") or "") == bpin]
        comp = sum(1 for h in hitos_p if str(col(h, "Estado") or "").upper() == "COMPLETADO")
        avance_fis = round((comp / len(hitos_p)) * 100) if hitos_p else 0
        avance_fin = round((total_pagos / total_aprop) * 100) if total_aprop > 0 else 0

        # Fuentes
        fuente_ppal = str(col(p, "Fuente Principal") or "RECURSOS_PROPIOS").upper()
        # Intentar construir fuentes desde ejecucion
        fuentes_dict = {}
        for e in ejec_p:
            fk = str(col(e, "Fuente", default="RECURSOS_PROPIOS")).upper()
            fuentes_dict[fk] = fuentes_dict.get(fk, 0) + to_int(col(e, "Apropiacion Final (COP)", "Apropiacion Inicial (COP)", default=0))
        if not fuentes_dict:
            fuentes_dict[fuente_ppal] = valor_total
        fuentes_js = "[" + ",".join(f'{{f:{s(k)},monto:{v}}}' for k, v in fuentes_dict.items()) + "]"

        # Contratos — tomar el primero como contrato principal (compatibilidad con app.js)
        contratos_p = [c for c in contratos_all if str(col(c, "BPIN") or "") == bpin]
        if contratos_p:
            c0 = contratos_p[0]
            contrato_js = (
                f'{{numero:{s(str(col(c0,"No. Contrato","")))},tipo:{s(str(col(c0,"Tipo Contrato","")))}'
                f',objeto:{s(str(col(c0,"Objeto","")))},contratista:{s(str(col(c0,"Contratista","")))}'
                f',nit:{s(str(col(c0,"NIT Contratista","")))},valor:{to_int(col(c0,"Valor (COP)",default=0))}'
                f',fecha:{s(str(col(c0,"Fecha Suscripcion","") or ""))}'
                f',estado:{s(str(col(c0,"Estado","") or ""))}'
                f',secopLink:{s(str(col(c0,"Link SECOP II","") or ""))}}}'
            )
        else:
            contrato_js = "null"

        # Hitos JS
        hitos_js_list = []
        for h in hitos_p:
            av = to_float(col(h, "Avance Fisico (%)", default=0))
            if av > 1:
                av = av / 100.0
            hitos_js_list.append(
                f'{{id:{s(str(col(h,"ID Hito","")))},descripcion:{s(str(col(h,"Descripcion Hito","")))}'
                f',fechaProg:{s(str(col(h,"Fecha Programada","")))},fechaReal:{s(str(col(h,"Fecha Real","") or ""))}'
                f',estado:{s(str(col(h,"Estado","PENDIENTE")))},avance:{round(av*100)}'
                f',responsable:{s(str(col(h,"Responsable","") or ""))}}}'
            )
        hitos_js = "[" + ",".join(hitos_js_list) + "]"

        items.append(
            f'  {{\n'
            f'    bpin:{s(bpin)}, nombre:{s(nombre)},\n'
            f'    sector:{s(sector)}, estado:{s(estado)},\n'
            f'    programaPDM:{s(prog_id)},\n'
            f'    fechaInicio:{s(fecha_i)}, fechaFin:{s(fecha_f)},\n'
            f'    descripcion:{s(desc)}, objetivo:{s(obj_g)},\n'
            f'    responsable:{s(resp)}, poblacionBeneficiada:{pob_ben},\n'
            f'    tipoPoblacion:{s(pob_tip)}, observaciones:{s(obs)},\n'
            f'    valorTotal:{valor_total},\n'
            f'    avanceFisico:{avance_fis}, avanceFinanciero:{avance_fin},\n'
            f'    fuentes:{fuentes_js},\n'
            f'    ejecucion:{ejec_js},\n'
            f'    contrato:{contrato_js},\n'
            f'    hitos:{hitos_js},\n'
            f'    codigoProductoDNP:{s(cod_dnp)},\n'
            f'    indicadorDNP:{s(ind_dnp)},\n'
            f'    unidadDNP:{s(uni_dnp)}\n'
            f'  }}'
        )

    return "let PROYECTOS = [\n" + ",\n".join(items) + "\n];"

# ── POLITICAS ─────────────────────────────────────────────────────────────────

def build_politicas():
    path = os.path.join(PLANT, "03_Politicas.xlsx")
    if not os.path.exists(path):
        print("  [AVISO] 03_Politicas.xlsx no encontrado")
        return "const POLITICAS = [];"

    politicas   = read_sheet(path, "Politicas")
    acciones_all = read_sheet(path, "PlanAccion")
    inds_all    = read_sheet(path, "Indicadores")
    print(f"  Políticas: {len(politicas)} | Acciones: {len(acciones_all)} | Indicadores: {len(inds_all)}")

    items = []
    for p in politicas:
        pid    = str(col(p, "ID Politica") or "")
        nombre = str(col(p, "Nombre") or "")
        sector = str(col(p, "Sector") or "GOBIERNO").upper()
        tipo   = str(col(p, "Tipo") or "")
        estado = str(col(p, "Estado") or "ACTIVA").upper()
        fecha_a = str(col(p, "Fecha Adopcion") or "")
        acto   = str(col(p, "Acto Administrativo") or "")
        entidad = str(col(p, "Entidad Lider") or "")
        desc   = str(col(p, "Descripcion") or "")
        pob_obj = str(col(p, "Poblacion Objetivo") or "")
        presup = to_int(col(p, "Presupuesto Total (COP)", default=0))

        # Acciones → planesAccion[0] (structure app.js expects)
        acciones_p = [a for a in acciones_all if str(col(a, "ID Politica") or "") == pid]
        acc_items = []
        total_avance = 0
        for a in acciones_p:
            av = to_float(col(a, "Avance (%)", default=0))
            if av > 1:
                av = av / 100.0
            av_pct = round(av * 100)
            total_avance += av_pct
            rec = to_int(col(a, "Recursos Asignados (COP)", default=0))
            est_raw = str(col(a, "Estado") or "PENDIENTE").upper()
            # Map to values app.js badge uses: COMPLETADA / EN_PROGRESO / PENDIENTE
            est_map = {"COMPLETADO": "COMPLETADA", "EN_PROGRESO": "EN_PROGRESO",
                       "PENDIENTE": "PENDIENTE", "CANCELADO": "PENDIENTE"}
            est_mapped = est_map.get(est_raw, est_raw)
            acc_items.append(
                f'{{nombre:{s(str(col(a,"Descripcion Accion","")))}'
                f',responsable:{s(str(col(a,"Responsable","") or ""))}'
                f',estado:{s(est_mapped)},avance:{av_pct}'
                f',recursos:{rec}}}'
            )

        avance_plan = round(total_avance / len(acciones_p)) if acciones_p else 0
        presup_plan = sum(to_int(col(a, "Recursos Asignados (COP)", default=0)) for a in acciones_p)
        acciones_js = "[" + ",".join(acc_items) + "]"
        plan_accion_js = (
            f'[{{nombre:"Plan de Acción {str(fecha_a)[:4] or 2024}"'
            f',presupuesto:{presup_plan},avance:{avance_plan},acciones:{acciones_js}}}]'
        )

        # Indicadores
        inds_p = [i for i in inds_all if str(col(i, "ID Politica") or "") == pid]
        inds_items = []
        for ind in inds_p:
            lb  = to_float(col(ind, "Linea Base", default=0))
            mt  = to_float(col(ind, "Meta Cuatrienio", default=0))
            log24 = to_float(col(ind, "Logro 2024", default=0))
            inds_items.append(
                f'{{nombre:{s(str(col(ind,"Nombre Indicador","")))}'
                f',unidad:{s(str(col(ind,"Unidad Medida","") or ""))}'
                f',lineaBase:{lb},metaCuatrienio:{mt},logrado2024:{log24}'
                f',fuenteVerificacion:{s(str(col(ind,"Fuente Verificacion","") or ""))}}}'
            )
        indicadores_js = "[" + ",".join(inds_items) + "]"

        items.append(
            f'  {{\n'
            f'    id:{s(pid)}, nombre:{s(nombre)},\n'
            f'    sector:{s(sector)}, tipo:{s(tipo)},\n'
            f'    estado:{s(estado)}, fechaAdopcion:{s(fecha_a)},\n'
            f'    actoAdministrativo:{s(acto)},\n'
            f'    entidadLider:{s(entidad)},\n'
            f'    descripcion:{s(desc)},\n'
            f'    poblacionObjetivo:{s(pob_obj)},\n'
            f'    presupuestoTotal:{presup},\n'
            f'    planesAccion:{plan_accion_js},\n'
            f'    indicadores:{indicadores_js}\n'
            f'  }}'
        )

    return "const POLITICAS = [\n" + ",\n".join(items) + "\n];"

# ── PLANES SECTORIALES ────────────────────────────────────────────────────────

def build_planes():
    path = os.path.join(PLANT, "04_PlanesSectoriales.xlsx")
    if not os.path.exists(path):
        print("  [AVISO] 04_PlanesSectoriales.xlsx no encontrado")
        return "const PLANES_SECTORIALES = [];"

    planes  = read_sheet(path, "PlanesSectoriales")
    metas_all = read_sheet(path, "ObjetivosMetas")
    print(f"  Planes: {len(planes)} | Metas: {len(metas_all)}")

    items = []
    for p in planes:
        pid    = str(col(p, "ID Plan") or "")
        nombre = str(col(p, "Nombre Plan") or "")
        sector = str(col(p, "Sector") or "GOBIERNO").upper()
        tipo   = str(col(p, "Tipo Plan") or "")
        estado = str(col(p, "Estado") or "ACTIVO").upper()
        vig_i_raw = col(p, "Vigencia Inicio") or ""
        vig_f_raw = col(p, "Vigencia Fin") or ""
        vig_i = str(vig_i_raw)[:10] if vig_i_raw else ""
        vig_f = str(vig_f_raw)[:10] if vig_f_raw else ""
        # Extract year from vigencia
        year_i = int(vig_i[:4]) if len(vig_i) >= 4 and vig_i[:4].isdigit() else 2024
        year_f = int(vig_f[:4]) if len(vig_f) >= 4 and vig_f[:4].isdigit() else 2027
        acto   = str(col(p, "Acto Administrativo") or "")
        presup_total = to_int(col(p, "Presupuesto Total (COP)", default=0))
        obj_g  = str(col(p, "Objetivo General") or "")
        resp   = str(col(p, "Responsable") or "")
        entidad = str(col(p, "Autoridad") or resp)

        # Distribute presupuesto by vigencia
        num_years = max(year_f - year_i + 1, 1)
        presup_per_year = round(presup_total / num_years)
        presup_dict_parts = ",".join(
            f'"{y}":{presup_per_year}' for y in range(year_i, year_f + 1)
        )
        presup_js = f'{{{presup_dict_parts}}}'

        # Metas / objetivos for this plan
        metas_p = [m for m in metas_all if str(col(m, "ID Plan") or "") == pid]
        obj_items = []
        avance_list = []
        for m in metas_p:
            lb  = to_float(col(m, "Linea Base", default=0))
            mt  = to_float(col(m, "Meta Total", default=0))
            log = to_float(col(m, "Logro Acumulado", default=0))
            pct = round(min((log / mt) * 100, 100)) if mt > 0 else 0
            avance_list.append(pct)
            obj_items.append(
                f'{{codigo:{s(str(col(m,"ID Meta","")))}'
                f',nombre:{s(str(col(m,"Descripcion Objetivo","") or ""))}'
                f',meta:{s(str(col(m,"Descripcion Meta","") or ""))}'
                f',indicador:{s(str(col(m,"Indicador","") or ""))}'
                f',esperado:{mt},logrado:{log},avance:{pct}}}'
            )
        avance_general = round(sum(avance_list) / len(avance_list)) if avance_list else 0
        objetivos_js = "[" + ",".join(obj_items) + "]"

        # avanceFisico / avanceFinanciero dicts by vigencia (usando el avance general para todos)
        av_fis_parts  = ",".join(f'"{y}":{avance_general}' for y in range(year_i, year_f + 1))
        av_fin_parts  = ",".join(f'"{y}":{max(avance_general-5,0)}' for y in range(year_i, year_f + 1))

        items.append(
            f'  {{\n'
            f'    id:{s(pid)}, nombre:{s(nombre)},\n'
            f'    sector:{s(sector)}, tipo:{s(tipo)},\n'
            f'    estado:{s(estado)},\n'
            f'    entidadLider:{s(entidad)}, responsable:{s(resp)},\n'
            f'    vigenciaInicio:{s(vig_i)}, vigenciaFin:{s(vig_f)},\n'
            f'    actoAdministrativo:{s(acto)},\n'
            f'    objetivoGeneral:{s(obj_g)},\n'
            f'    presupuesto:{presup_js},\n'
            f'    avanceFisico:{{{av_fis_parts}}},\n'
            f'    avanceFinanciero:{{{av_fin_parts}}},\n'
            f'    objetivos:{objetivos_js}\n'
            f'  }}'
        )

    return "const PLANES_SECTORIALES = [\n" + ",\n".join(items) + "\n];"

# ── main ──────────────────────────────────────────────────────────────────────

def main():
    print("Leyendo plantillas Excel...")
    static = read_static_header()

    pdm_js      = build_pdm()
    proy_js     = build_proyectos()
    pol_js      = build_politicas()
    planes_js   = build_planes()

    # Utility functions block (preserved always)
    utility_js = r"""
// ─── Funciones utilitarias ────────────────────────────────────────────────────

function formatBPIN(bpin) {
  if (!bpin) return '';
  return String(bpin).replace(/\D/g,'').padStart(14,'0');
}

function formatCOP(v) {
  const n = Number(v)||0;
  if (n >= 1e12) return '$'+(n/1e12).toFixed(1)+'B';
  if (n >= 1e9)  return '$'+(n/1e9).toFixed(1)+'B';
  if (n >= 1e6)  return '$'+(n/1e6).toFixed(0)+'M';
  if (n >= 1e3)  return '$'+(n/1e3).toFixed(0)+'K';
  return '$'+n.toFixed(0);
}

function formatCOPFull(v) {
  return new Intl.NumberFormat('es-CO',{style:'currency',currency:'COP',minimumFractionDigits:0}).format(Number(v)||0);
}

function semaforoColor(fisico, financiero, estado) {
  if (estado==='SUSPENDIDO') return 'ROJO';
  if (estado==='TERMINADO'||estado==='CERRADO') return 'VERDE';
  if (estado==='FORMULACION'||estado==='VIABILIZADO'||estado==='REGISTRADO') return null;
  const f=Number(fisico)||0, g=Number(financiero)||0;
  const brecha=Math.abs(f-g);
  if (brecha>=25||(f<15&&g>35)) return 'ROJO';
  if (brecha>=15) return 'AMARILLO';
  return 'VERDE';
}

const TENDENCIA_MENSUAL = [
  {mes:'Ene',apropiacion:7.1,pagos:0.3,fisico:2},
  {mes:'Feb',apropiacion:7.1,pagos:1.8,fisico:6},
  {mes:'Mar',apropiacion:7.1,pagos:4.2,fisico:12},
  {mes:'Abr',apropiacion:7.1,pagos:7.9,fisico:19},
  {mes:'May',apropiacion:7.1,pagos:12.1,fisico:27},
  {mes:'Jun',apropiacion:7.1,pagos:18.4,fisico:36},
  {mes:'Jul',apropiacion:7.1,pagos:24.7,fisico:44},
  {mes:'Ago',apropiacion:7.1,pagos:31.2,fisico:52},
  {mes:'Sep',apropiacion:7.1,pagos:38.6,fisico:61},
  {mes:'Oct',apropiacion:7.1,pagos:46.3,fisico:70},
  {mes:'Nov',apropiacion:7.1,pagos:54.1,fisico:79},
  {mes:'Dic',apropiacion:7.1,pagos:61.8,fisico:87},
];
"""

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    out = (
        static.rstrip() + "\n\n"
        f"// ─── Datos generados por cargar_datos.py — {ts} ───────────────────────────\n\n"
        + pdm_js + "\n\n"
        + proy_js + "\n\n"
        + pol_js + "\n\n"
        + planes_js + "\n"
        + utility_js
    )

    with open(OUT_JS, "w", encoding="utf-8") as f:
        f.write(out)

    print(f"\ndata.js actualizado: {OUT_JS}")
    print("Recarga el navegador (F5) para ver los cambios.")


if __name__ == "__main__":
    main()
