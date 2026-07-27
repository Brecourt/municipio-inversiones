"""
cargar_poai.py — Genera js/data.js desde el archivo maestro de seguimiento.

FUENTE UNICA DE VERDAD:
  plantillas/POAI_SEGUIMIENTO_2026.xlsx
  (POAI + Plan de Accion + seguimiento fisico y financiero del mes)

Para actualizar cada mes:
  1. Reemplace plantillas/POAI_SEGUIMIENTO_2026.xlsx con el archivo nuevo
  2. Ejecute:  python cargar_poai.py
  3. El script detecta solo la fecha de corte desde los encabezados

El mapeo de columnas es POR NOMBRE (no por posicion), de modo que agregar o
quitar columnas en el Excel no rompe la carga.

Complementa con:
  plantillas/PDT_05284.xlsx          (Plan de Desarrollo Territorial)
  plantillas/05_IndicadoresPDM.xlsx  (indicadores radar)
  plantillas/02_Proyectos.xlsx       (hoja Contratos)

Req:  pip install pandas openpyxl
"""
import os, re, csv, io, unicodedata
import pandas as pd
from datetime import date, datetime
from openpyxl import load_workbook

BASE   = os.path.dirname(os.path.abspath(__file__))
PLANT  = os.path.join(BASE, "plantillas")
OUT_JS = os.path.join(BASE, "js", "data.js")

POAI_XLS  = os.path.join(PLANT, "POAI_SEGUIMIENTO_2026.xlsx")
CONTR_CSV = os.path.join(PLANT, "CONTRATACION_2026.csv")
PDT_FILE  = os.path.join(PLANT, "PDT_05284.xlsx")
RADAR_XLS = os.path.join(PLANT, "05_IndicadoresPDM.xlsx")
PROY_XLS  = os.path.join(PLANT, "02_Proyectos.xlsx")

VIGENCIA = 2026

# ── Sector desde codigo de programa MGA ──────────────────────
SECTOR_MAP = {
    1905: 'SALUD Y PROTECCION SOCIAL',    1906: 'SALUD Y PROTECCION SOCIAL',
    2201: 'EDUCACION',
    3301: 'CULTURA',                      3302: 'CULTURA',
    3502: 'COMERCIO, INDUSTRIA Y TURISMO',
    4001: 'VIVIENDA, CIUDAD Y TERRITORIO',
    4003: 'AGUA',
    4101: 'INCLUSION SOCIAL Y RECONCILIACION', 4102: 'INCLUSION SOCIAL Y RECONCILIACION',
    4103: 'INCLUSION SOCIAL Y RECONCILIACION', 4104: 'INCLUSION SOCIAL Y RECONCILIACION',
    4301: 'DEPORTE Y RECREACION',
    3201: 'AMBIENTE Y DESARROLLO SOSTENIBLE', 3202: 'AMBIENTE Y DESARROLLO SOSTENIBLE',
    3203: 'AMBIENTE Y DESARROLLO SOSTENIBLE', 3208: 'AMBIENTE Y DESARROLLO SOSTENIBLE',
    2102: 'MINAS Y ENERGIA',               2104: 'MINAS Y ENERGIA',
    2402: 'TRANSPORTE',                    2409: 'TRANSPORTE',
    1702: 'AGRICULTURA Y DESARROLLO RURAL',1704: 'AGRICULTURA Y DESARROLLO RURAL',
    1709: 'AGRICULTURA Y DESARROLLO RURAL',
    4502: 'GOBIERNO', 4503: 'GOBIERNO', 4599: 'GOBIERNO',
    1202: 'JUSTICIA Y DEL DERECHO', 1206: 'JUSTICIA Y DEL DERECHO',
}

# Fuentes de financiacion: nombre normalizado de columna → clave en data.js
FUENTES_MAP = [
    ('recursos propios',            'RECURSOS_PROPIOS'),
    ('sgp alimentacion escolar',    'SGP_PAE'),
    ('sgp apsb',                    'SGP_APSB'),
    ('sgp cultura',                 'SGP_CULTURA'),
    ('sgp deporte',                 'SGP_DEPORTE'),
    ('sgp educacion',               'SGP_EDUCACION'),
    ('sgp libre destinacion',       'SGP_LD'),
    ('sgp libre inversion',         'SGP_LI'),
    ('sgp salud',                   'SGP_SALUD'),
    ('regalias',                    'SGR'),
    ('cofinanciacion departamento', 'COFINANCIACION'),
    ('cofinanciacion nacion',       'COFINANCIACION'),
    ('credito',                     'CREDITO'),
    ('otros',                       'RECURSOS_PROPIOS'),
]

MESES = {
    'enero':1,'febrero':2,'marzo':3,'abril':4,'mayo':5,'junio':6,
    'julio':7,'agosto':8,'septiembre':9,'setiembre':9,'octubre':10,
    'noviembre':11,'diciembre':12,
}
MESES_NOM = ['','Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio',
             'Agosto','Septiembre','Octubre','Noviembre','Diciembre']

# ── Helpers ──────────────────────────────────────────────────
def norm(t):
    """Normaliza texto: sin acentos, sin simbolos, minusculas, espacios simples."""
    t = unicodedata.normalize('NFKD', str(t))
    t = ''.join(c for c in t if not unicodedata.combining(c))
    t = t.replace('�', '').lower()
    return re.sub(r'[^a-z0-9]+', ' ', t).strip()

def s(v):
    if v is None:
        return '""'
    v = str(v).replace("\\", "\\\\").replace('"', '\\"')
    v = v.replace("\r", " ").replace("\n", " ")
    return '"' + re.sub(r'\s+', ' ', v).strip() + '"'

def to_int(v, default=0):
    try:
        f = float(str(v).replace(",", "").replace("$", "").strip())
        return default if f != f else int(round(f))
    except Exception:
        return default

def to_float(v, default=0.0):
    try:
        f = float(str(v).replace(",", "").replace("$", "").replace("%", "").strip())
        return default if f != f else f
    except Exception:
        return default

def pct(v):
    """Convierte a porcentaje 0-N. Acepta fraccion (0.35) o porcentaje (35)."""
    f = to_float(v, 0.0)
    if f == 0:
        return 0.0
    # Los archivos SisPT traen fracciones; >1.0 se asume ya en porcentaje solo si es muy grande
    return round(f * 100, 1) if f <= 1.5 else round(f, 1)

class ColFinder:
    """Localiza columnas por nombre normalizado, inmune a cambios de posicion."""
    def __init__(self, columns):
        self.cols = list(columns)
        self.by_norm = {}
        for c in self.cols:
            self.by_norm.setdefault(norm(c), c)

    def exact(self, *names):
        for nm in names:
            if nm in self.by_norm:
                return self.by_norm[nm]
        return None

    def starts(self, *prefixes):
        for p in prefixes:
            for nc, orig in self.by_norm.items():
                if nc.startswith(p):
                    return orig
        return None

    def contains(self, *kws, exclude=()):
        for k in kws:
            for nc, orig in self.by_norm.items():
                if k in nc and not any(x in nc for x in exclude):
                    return orig
        return None

def detectar_corte(col_pagos):
    """Extrae la fecha de corte del encabezado, ej: '... - 30 de Junio' → (30,6)."""
    n = norm(col_pagos)
    m = re.search(r'(\d{1,2})\s*(?:de\s*)?(' + '|'.join(MESES) + r')', n)
    if m:
        return int(m.group(1)), MESES[m.group(2)]
    m = re.search(r'(' + '|'.join(MESES) + r')', n)
    if m:
        return 30, MESES[m.group(1)]
    return None, None

# ── Lectura de hojas auxiliares ──────────────────────────────
def read_sheet_rows(path, sheet_name):
    if not os.path.exists(path):
        return []
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    rows = list(wb[sheet_name].iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
    out = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        d = {}
        for h, v in zip(headers, row):
            if isinstance(v, (date, datetime)):
                v = v.strftime("%Y-%m-%d") if isinstance(v, datetime) else v.isoformat()
            d[h] = v
        out.append(d)
    return out

def col_val(row, *keys, default=None):
    for k in keys:
        v = row.get(k)
        if v is not None and str(v).strip() not in ("", "None", "nan"):
            return v
    return default

# ── Lectura del POAI de seguimiento ──────────────────────────
def read_poai():
    df = pd.read_excel(POAI_XLS, sheet_name=0, header=1, dtype={'BPIM': str})
    df.columns = [str(c).strip() for c in df.columns]
    f = ColFinder(df.columns)

    cols = {
        'linea':       f.exact('linea'),
        'componente':  f.exact('componente'),
        'cod_prog':    f.exact('cod programa'),
        'programa':    f.exact('programa'),
        'prod_mga':    f.exact('producto mga'),
        'cod_ind':     f.exact('cod indicador de producto'),
        'ind_mga':     f.exact('indicador mga'),
        'desc_pdm':    f.exact('descripcion indicador pdm'),
        'unidad':      f.exact('unidad de medida'),
        'meta_cuat':   f.exact('meta cuatrienio'),
        'meta_vig':    f.exact(f'meta {VIGENCIA}'),
        'av_fis':      f.exact(f'avance fisico {VIGENCIA}'),
        'pct_fis':     f.exact('avance fisico'),
        'bpin':        f.exact('bpim', 'bpin'),
        'nombre':      f.exact('nombre del proyecto programa'),
        'dependencia': f.exact('dependencia responsable'),
        'actividades': f.exact('actividades acciones estrategias'),
        'total':       f.exact('total programado') or f.contains('total programado'),
        'compromisos': f.starts('avance financiero comrpomisos', 'avance financiero compromisos'),
        'obligaciones':f.starts('avance financiero obligaciones'),
        'pagos':       f.starts('avance financiero pagos'),
        'pct_fin':     f.exact('avance financiero'),
    }

    faltantes = [k for k, v in cols.items() if v is None and k not in ('unidad',)]
    if faltantes:
        raise SystemExit(
            "[ERROR] No se encontraron estas columnas en el Excel: "
            + ", ".join(faltantes)
            + "\n        Columnas disponibles: " + ", ".join(map(str, df.columns))
        )

    # Fuentes de financiacion presentes
    fuentes = []
    for nombre_norm, clave in FUENTES_MAP:
        c = f.exact(nombre_norm)
        if c is not None:
            fuentes.append((c, clave))

    # Filtrar filas sin BPIN
    df = df[df['BPIM'].notna()]
    df = df[~df['BPIM'].astype(str).str.strip().str.lower().isin(['', 'nan', 'none'])]

    dd, mm = detectar_corte(cols['pagos'])
    return df, cols, fuentes, (dd, mm)

# ── Contratacion (reporte a entes de control) ────────────────
def _fecha_iso(v):
    """Convierte DD/MM/AAAA a AAAA-MM-DD. Descarta placeholders (0/01/1900)."""
    t = str(v or '').strip()
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', t)
    if not m:
        return ''
    d, mth, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if y < 1950 or d == 0 or mth == 0:
        return ''
    return f"{y:04d}-{mth:02d}-{d:02d}"

def read_contratos(bpins_validos):
    """
    Lee el consolidado de contratacion (delimitado por |).
    El campo 'proyectos' puede traer varios BPIN separados por ; o ,
    Devuelve (lista_de_contratos, indice_por_bpin, avisos)
    """
    if not os.path.exists(CONTR_CSV):
        return [], {}, ["No se encontro plantillas/CONTRATACION_2026.csv"]

    raw = open(CONTR_CSV, 'rb').read()
    txt = None
    for enc in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
        try:
            txt = raw.decode(enc); break
        except Exception:
            continue
    if txt is None:
        return [], {}, ["No se pudo decodificar CONTRATACION_2026.csv"]

    filas = list(csv.DictReader(io.StringIO(txt), delimiter='|'))
    contratos, por_bpin, avisos = [], {}, []

    for i, r in enumerate(filas, start=2):
        g = lambda k: str(r.get(k, '') or '').strip()

        crudo = g('proyectos')
        bpins = [b.strip() for b in re.split(r'[;,]', crudo) if b.strip()]
        validos    = [b for b in bpins if b in bpins_validos]
        invalidos  = [b for b in bpins if b not in bpins_validos]

        numero = g('numeroContratoInternoConvenio') or g('idContratoSecop')
        if invalidos:
            avisos.append(
                f"fila {i}: contrato {numero} referencia BPIN no registrado en el POAI: "
                + ", ".join(invalidos)
            )

        c = {
            'numero':      numero,
            'idSecop':     g('idContratoSecop'),
            'tipo':        g('tipoContrato'),
            'modalidad':   g('modalidad'),
            'objeto':      g('objetoContrato'),
            'contratista': g('nombreContratista'),
            'tipoDoc':     g('tipoIdentificacionContratista'),
            'nit':         g('identificacionContratista'),
            'valor':       to_int(g('compromisos')),
            'adiciones':   to_int(g('valorTotalAdiciones')),
            'nAdiciones':  to_int(g('numeroAdiciones')),
            'nProrrogas':  to_int(g('numeroProrrogas')),
            'nSuspension': to_int(g('numeroSuspenciones')),
            'plazoDias':   to_int(g('plazoContrato')),
            'fecha':       _fecha_iso(g('fechaFirmaContrato')),
            'fechaInicio': _fecha_iso(g('fechaInicioContrato')),
            'fechaFin':    _fecha_iso(g('fechaTerminaContrato')),
            'fechaActa':   _fecha_iso(g('fechaActaLiquidacion')),
            'estado':      g('estadoContrato'),
            'linea':       g('lineaProyecto'),
            'fSGP':        to_int(g('valorTotalSGP')),
            'fSGR':        to_int(g('valorTotalSGR')),
            'fPropios':    to_int(g('valorTotalPropios')),
            'fNacion':     to_int(g('valorTotalPgn')),
            'fOtros':      to_int(g('valorTotalOtros')),
            'observaciones': g('observaciones'),
            'bpins':       validos,
            'bpinsCrudo':  crudo,
            'enlazado':    bool(validos),
            'compartido':  len(validos) > 1,
        }
        contratos.append(c)
        for b in validos:
            por_bpin.setdefault(b, []).append(c)

    return contratos, por_bpin, avisos

def contrato_a_js(c):
    campos = [
        ('numero', s), ('idSecop', s), ('tipo', s), ('modalidad', s), ('objeto', s),
        ('contratista', s), ('tipoDoc', s), ('nit', s),
        ('valor', str), ('adiciones', str), ('nAdiciones', str), ('nProrrogas', str),
        ('nSuspension', str), ('plazoDias', str),
        ('fecha', s), ('fechaInicio', s), ('fechaFin', s), ('fechaActa', s),
        ('estado', s), ('linea', s),
        ('fSGP', str), ('fSGR', str), ('fPropios', str), ('fNacion', str), ('fOtros', str),
        ('observaciones', s),
    ]
    partes = [f'{k}:{fn(c[k])}' for k, fn in campos]
    partes.append('bpins:[' + ",".join(s(b) for b in c['bpins']) + ']')
    partes.append('bpin:' + s(c['bpins'][0] if c['bpins'] else ''))
    partes.append('bpinsCrudo:' + s(c['bpinsCrudo']))
    partes.append('enlazado:' + ('true' if c['enlazado'] else 'false'))
    partes.append('compartido:' + ('true' if c['compartido'] else 'false'))
    partes.append('secopLink:' + s(
        'https://www.contratos.gov.co/consultas/detalleProceso.do?numConstancia=' + c['idSecop']
        if c['idSecop'] else ''
    ))
    return '{' + ",".join(partes) + '}'

# ── PDM ──────────────────────────────────────────────────────
def read_radar():
    if not os.path.exists(RADAR_XLS):
        return "[]"
    wb = load_workbook(RADAR_XLS, data_only=True)
    if "IndicadoresRadar" not in wb.sheetnames:
        return "[]"
    rows = list(wb["IndicadoresRadar"].iter_rows(values_only=True))
    if not rows:
        return "[]"
    headers = [str(h).strip() if h else f"c{i}" for i, h in enumerate(rows[0])]
    items = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        d = dict(zip(headers, row))
        eid = str(d.get("ID Eje", "") or "")
        if not eid or eid.startswith("*") or eid.startswith("⚠"):
            continue
        label = str(d.get("Nombre Eje (etiqueta corta)", eid) or eid)
        items.append(f'{{eje:{s(label)},logrado:{to_float(d.get("Logrado (%)",0))},'
                     f'meta:{to_float(d.get("Meta (%)",100))}}}')
    return "[" + ",".join(items) + "]" if items else "[]"

def build_pdm(df, cols, avance_global_fis, avance_global_fin):
    LINEAS = {
        'LE-1': 'FRONTINO NOS UNE CON TEJIDO SOCIAL, INCLUSION Y RECONCILIACION.',
        'LE-2': 'FRONTINO NOS UNE CON SOSTENIBILIDAD AMBIENTAL Y DESARROLLO SUSTENTABLE.',
        'LE-3': 'FRONTINO NOS UNE CON COMPETITIVIDAD Y DESARROLLO ECONOMICO LOCAL.',
        'LE-4': 'FRONTINO NOS UNE CON BUEN GOBIERNO, GOBERNANZA Y GOBERNABILIDAD.',
    }
    try:
        lineas_raw = pd.read_excel(PDT_FILE, sheet_name='l\xedneas estrat\xe9gicas')
        for _, r in lineas_raw.iterrows():
            lid = str(r.iloc[0]).strip()
            if lid in LINEAS:
                LINEAS[lid] = str(r.iloc[2]).strip()
    except Exception:
        pass

    ejes = {lid: {'id': lid, 'nombre': nm, 'programas': []} for lid, nm in LINEAS.items()}
    for _, r in df.iterrows():
        linea = str(r.get(cols['linea'], '') or '').strip()
        if linea not in ejes:
            linea = 'LE-1'
        try:
            raw = r.get(cols['cod_prog'], '')
            prog_cod = str(int(float(str(raw)))) if str(raw) not in ('nan', '', 'None') else ''
        except Exception:
            prog_cod = ''
        ind = str(r.get(cols['ind_mga'], '') or '').strip()
        meta4 = str(r.get(cols['meta_cuat'], '') or '').strip()
        ejes[linea]['programas'].append(
            f'{{id:{s(prog_cod)},nombre:{s(str(r.get(cols["programa"],"") or "").strip())},'
            f'meta:{s((meta4 + " " + ind).strip())},indicador:{s(ind)},lineaBase:0,'
            f'presupuesto:{to_int(r.get(cols["total"], 0))}}}'
        )

    # Radar: usa el archivo si existe; si no, calcula por eje desde el seguimiento
    radar = read_radar()
    if radar == "[]":
        rad = []
        for lid, e in ejes.items():
            sub = df[df[cols['linea']].astype(str).str.strip() == lid]
            if len(sub) == 0:
                continue
            log = sum(min(pct(v), 100) for v in sub[cols['pct_fis']]) / len(sub)
            rad.append(f'{{eje:{s(lid)},logrado:{round(log,1)},meta:100}}')
        radar = "[" + ",".join(rad) + "]"

    ejes_js = ",".join(
        f'{{id:{s(e["id"])},nombre:{s(e["nombre"])},programas:[{",".join(e["programas"])}]}}'
        for e in ejes.values()
    )
    return (
        'const PDM = {\n'
        f'  municipio:{s("Frontino")},\n'
        f'  periodo:{s("2024-2027")},\n'
        f'  nombre:{s(chr(34) + "Frontino Nos Une" + chr(34) + " 2024-2027")},\n'
        f'  alcalde:{s("Luz Gabriela Rivera Cano")},\n'
        f'  acuerdo:{s("Acuerdo 007 de 31 de Mayo de 2024")},\n'
        f'  inversionTotal:{to_int(df[cols["total"]].sum())},\n'
        f'  avanceFisicoGlobal:{avance_global_fis},\n'
        f'  avanceFinancieroGlobal:{avance_global_fin},\n'
        f'  indicadoresRadar:{radar},\n'
        f'  ejes:[{ejes_js}]\n'
        '};'
    )

# ── PROYECTOS ────────────────────────────────────────────────
def build_proyectos(df, cols, fuentes, cont_idx):
    items = []
    for _, r in df.iterrows():
        bpin  = str(r.get(cols['bpin'], '') or '').strip()
        linea = str(r.get(cols['linea'], '') or '').strip()
        try:
            raw = r.get(cols['cod_prog'], '')
            prog = str(int(float(str(raw)))) if str(raw) not in ('nan', '', 'None') else ''
        except Exception:
            prog = ''
        prog_int = int(prog) if prog.isdigit() else 0
        sector = SECTOR_MAP.get(prog_int, 'GOBIERNO')

        ind_mga = str(r.get(cols['ind_mga'], '') or '').strip()
        nombre  = str(r.get(cols['nombre'], '') or '').strip()
        nombre  = re.sub(r'\s+', ' ', nombre) or ind_mga

        total  = to_int(r.get(cols['total'], 0))
        comp   = to_int(r.get(cols['compromisos'], 0))
        obl    = to_int(r.get(cols['obligaciones'], 0))
        pagos  = to_int(r.get(cols['pagos'], 0))
        av_fis = pct(r.get(cols['pct_fis'], 0))
        av_fin = pct(r.get(cols['pct_fin'], 0))

        # Estado real segun el ciclo presupuestal
        if total > 0 and pagos >= total * 0.99:
            estado = 'TERMINADO'
        elif comp > 0 or obl > 0 or pagos > 0:
            estado = 'EN_EJECUCION'
        else:
            estado = 'REGISTRADO'

        fdict = {}
        for col, clave in fuentes:
            v = to_int(r.get(col, 0))
            if v > 0:
                fdict[clave] = fdict.get(clave, 0) + v
        if not fdict and total > 0:
            fdict['RECURSOS_PROPIOS'] = total
        fuentes_js = "[" + ",".join(f'{{f:{s(k)},monto:{v}}}' for k, v in fdict.items()) + "]"

        ejec_js = (f'[{{vigencia:{VIGENCIA},apropiacion:{total},cdp:{comp},rp:{comp},'
                   f'obligaciones:{obl},pagos:{pagos},fuente:""}}]')

        cs = cont_idx.get(bpin, [])
        c_uno   = contrato_a_js(cs[0]) if cs else "null"
        c_todos = "[" + ",".join(contrato_a_js(c) for c in cs) + "]"
        val_contratado = sum(c['valor'] for c in cs)

        items.append(
            '  {\n'
            f'    bpin:{s(bpin)}, nombre:{s(nombre)},\n'
            f'    sector:{s(sector)}, estado:{s(estado)},\n'
            f'    programaPDM:{s(prog)},\n'
            f'    fechaInicio:"", fechaFin:"",\n'
            f'    descripcion:{s(str(r.get(cols["desc_pdm"],"") or "").strip())},\n'
            f'    objetivo:{s(ind_mga)},\n'
            f'    responsable:{s(str(r.get(cols["dependencia"],"") or "").strip())},\n'
            f'    actividades:{s(str(r.get(cols["actividades"],"") or "").strip())},\n'
            f'    poblacionBeneficiada:0, tipoPoblacion:"", observaciones:"",\n'
            f'    valorTotal:{total},\n'
            f'    compromisos:{comp}, obligaciones:{obl}, pagos:{pagos},\n'
            f'    avanceFisico:{av_fis}, avanceFinanciero:{av_fin},\n'
            f'    fuentes:{fuentes_js},\n'
            f'    ejecucion:{ejec_js},\n'
            f'    contrato:{c_uno},\n'
            f'    contratos:{c_todos},\n'
            f'    valorContratado:{val_contratado}, numContratos:{len(cs)},\n'
            f'    hitos:[],\n'
            f'    codigoProductoDNP:{s(str(r.get(cols["prod_mga"],"") or "").strip().rstrip(".0"))},\n'
            f'    indicadorDNP:{s(str(r.get(cols["cod_ind"],"") or "").strip().rstrip(".0"))},\n'
            f'    unidadDNP:{s(str(r.get(cols["unidad"],"") or "").strip() if cols.get("unidad") else "")},\n'
            f'    productoNombre:{s(ind_mga)},\n'
            f'    metaCuatrienio:{s(str(r.get(cols["meta_cuat"],"") or "").strip())},\n'
            f'    metaVigencia:{s(str(r.get(cols["meta_vig"],"") or "").strip())},\n'
            f'    avanceMeta:{s(str(r.get(cols["av_fis"],"") or "").strip())},\n'
            f'    lineaEstrategica:{s(linea)},\n'
            f'    componente:{s(str(r.get(cols["componente"],"") or "").strip())}\n'
            '  }'
        )
    return "let PROYECTOS = [\n" + ",\n".join(items) + "\n];"

# ── Cabecera estatica ────────────────────────────────────────
def read_static_header():
    try:
        with open(OUT_JS, "r", encoding="utf-8") as f:
            content = f.read()
        for marker in [r'\n// ─+ Datos generados', r'\n// ─+ Seguimiento',
                       r'\nconst CORTE\s*=', r'\nconst PDM\s*=', r'\nlet PROYECTOS\s*=']:
            m = re.search(marker, content)
            if m:
                return content[:m.start()].rstrip()
        return content
    except Exception:
        return "// data.js generado por cargar_poai.py"

UTILITY_JS = r"""
function semaforoColor(fisico, financiero) {
  const f = fisico || 0, g = financiero || 0;
  if (f === 0 && g === 0) return 'ROJO';
  const brecha = Math.abs(f - g);
  if (f === 0 || g === 0) return 'AMARILLO';
  if (brecha <= 10) return 'VERDE';
  if (brecha <= 25) return 'AMARILLO';
  return 'ROJO';
}
// Escala corta en español: 1e9 son miles de millones (mmM), no "billones".
// La B se reserva para el billón real (1e12), como en la escala larga que usa Colombia.
function formatCOP(v) {
  if (!v && v !== 0) return '$0';
  const abs = Math.abs(v);
  if (abs >= 1e12) return '$' + (v/1e12).toFixed(1) + ' B';
  if (abs >= 1e9)  return '$' + (v/1e9).toFixed(1)  + ' mmM';
  if (abs >= 1e6)  return '$' + (v/1e6).toFixed(1)  + ' M';
  if (abs >= 1e3)  return '$' + (v/1e3).toFixed(0)  + ' K';
  return '$' + v.toFixed(0);
}
// Valor completo en pesos, con separador de miles: $1.234.567.890
function formatCOPFull(v) {
  const n = Number(v) || 0;
  return '$' + n.toLocaleString('es-CO', { maximumFractionDigits: 0 });
}
// El BPIN se muestra tal cual, solo digitos y sin separadores
function formatBPIN(v) {
  if (!v) return '—';
  return String(v).replace(/\D/g, '');
}
const POLITICAS = [];
const PLANES_SECTORIALES = [];
"""

# ── MAIN ─────────────────────────────────────────────────────
def main():
    print("=== cargar_poai.py ===")
    print(f"\n[1] Fuente unica: {os.path.basename(POAI_XLS)}")
    df, cols, fuentes, (dd, mm) = read_poai()

    total = to_int(df[cols['total']].sum())
    comp  = to_int(df[cols['compromisos']].sum())
    obl   = to_int(df[cols['obligaciones']].sum())
    pag   = to_int(df[cols['pagos']].sum())

    n = len(df)
    fis_prom = round(sum(min(pct(v), 100) for v in df[cols['pct_fis']]) / n, 1) if n else 0
    fin_glob = round(pag / total * 100, 1) if total else 0

    corte_txt = f"{dd} de {MESES_NOM[mm]} de {VIGENCIA}" if mm else f"vigencia {VIGENCIA}"
    corte_iso = f"{VIGENCIA}-{mm:02d}-{dd:02d}" if mm else f"{VIGENCIA}-12-31"

    print(f"    Proyectos:      {n}  ({df[cols['bpin']].nunique()} BPIN unicos)")
    print(f"    Corte:          {corte_txt}")
    print(f"    Programado:     ${total:,}")
    print(f"    Compromisos:    ${comp:,}  ({comp/total*100:.1f}%)" if total else "")
    print(f"    Obligaciones:   ${obl:,}  ({obl/total*100:.1f}%)" if total else "")
    print(f"    Pagos:          ${pag:,}  ({fin_glob}%)")
    print(f"    Avance fisico:  {fis_prom}% promedio")

    print(f"\n[2] Contratacion: {os.path.basename(CONTR_CSV)}")
    bpins_validos = set(str(b).strip() for b in df[cols['bpin']])
    contratos, cont_idx, avisos = read_contratos(bpins_validos)
    val_contr = sum(c['valor'] for c in contratos)
    enlazados = sum(1 for c in contratos if c['enlazado'])
    print(f"    Contratos:      {len(contratos)}")
    print(f"    Valor total:    ${val_contr:,}")
    print(f"    Enlazados:      {enlazados}/{len(contratos)}  ({len(cont_idx)} proyectos con contrato)")
    compartidos = [c for c in contratos if c['compartido']]
    if compartidos:
        print(f"    Compartidos:    {len(compartidos)} contrato(s) cubren varios proyectos")
        for c in compartidos:
            print(f"                    {c['numero']} -> {len(c['bpins'])} BPIN")
    for a in avisos:
        print(f"    [!] {a}")

    print(f"\n[3] Construyendo PDM...")
    pdm_js = build_pdm(df, cols, fis_prom, fin_glob)

    print(f"[4] Construyendo PROYECTOS...")
    proy_js = build_proyectos(df, cols, fuentes, cont_idx)

    contratos_js = ("const CONTRATOS = [\n  "
                    + ",\n  ".join(contrato_a_js(c) for c in contratos)
                    + "\n];")

    corte_js = (
        'const CORTE = {\n'
        f'  fecha:{s(corte_iso)},\n'
        f'  etiqueta:{s(corte_txt)},\n'
        f'  vigencia:{VIGENCIA},\n'
        f'  programado:{total},\n'
        f'  compromisos:{comp},\n'
        f'  obligaciones:{obl},\n'
        f'  pagos:{pag},\n'
        f'  numContratos:{len(contratos)},\n'
        f'  valorContratado:{val_contr}\n'
        '};'
    )

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    out = (
        read_static_header() + "\n\n"
        f"// ─── Seguimiento fisico y financiero al {corte_txt} ───\n"
        f"// ─── Fuentes: {os.path.basename(POAI_XLS)} + {os.path.basename(CONTR_CSV)} ───\n"
        f"// ─── Generado {ts} por cargar_poai.py ───\n\n"
        + corte_js + "\n\n" + pdm_js + "\n\n" + proy_js + "\n\n"
        + contratos_js + "\n\n" + UTILITY_JS
    )
    # Guarda: detectar literales de Python que se hayan colado al JS.
    # Producen un ReferenceError en el navegador que la validacion de
    # sintaxis no alcanza a ver, dejando la pagina en blanco.
    fugas = []
    for pat, desc in [
        (r'[:\[,]\s*True\b',   'booleano True de Python'),
        (r'[:\[,]\s*False\b',  'booleano False de Python'),
        (r'[:\[,]\s*None\b',   'None de Python'),
        (r"\{'[a-zA-Z_]+':",   'diccionario de Python serializado con str()'),
    ]:
        for m in re.finditer(pat, out):
            linea = out.count('\n', 0, m.start()) + 1
            fugas.append(f"linea {linea}: {desc} -> {out[m.start():m.start()+60]!r}")
    if fugas:
        print("\n[ERROR] Se filtraron literales de Python al JavaScript generado:")
        for f_ in fugas[:8]:
            print("   " + f_)
        print("   No se escribio data.js. Revise cargar_poai.py.")
        raise SystemExit(1)

    with open(OUT_JS, "w", encoding="utf-8") as f:
        f.write(out)

    print(f"\n[OK] {OUT_JS}")
    print(f"     {len(df)} proyectos · {len(contratos)} contratos · corte {corte_txt}")

if __name__ == "__main__":
    main()
